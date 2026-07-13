import os, json, re, random, datetime, time
from functools import wraps
from flask import (Flask, render_template, request, jsonify,
                   redirect, url_for, session, send_from_directory,
                   Response, send_file)
import io, csv
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
from flask_socketio import SocketIO, join_room, leave_room, emit
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from datetime import datetime, timedelta

# ─────────────────────────────────────────────
# JWT — lightweight pure-Python implementation
# (no extra pip install needed)
# ─────────────────────────────────────────────
import hmac, hashlib, base64

def _b64url(data: bytes) -> str:
    return base64.urlsafe_b64encode(data).rstrip(b'=').decode()

def _b64url_decode(s: str) -> bytes:
    pad = 4 - len(s) % 4
    return base64.urlsafe_b64decode(s + '=' * (pad % 4))

def jwt_encode(payload: dict, secret: str, expires_in: int = 3600) -> str:
    """Create a signed JWT. expires_in = seconds."""
    header  = _b64url(json.dumps({"alg":"HS256","typ":"JWT"}).encode())
    payload = dict(payload)
    payload['exp'] = int(time.time()) + expires_in
    payload_b64 = _b64url(json.dumps(payload).encode())
    sig = hmac.new(secret.encode(), f"{header}.{payload_b64}".encode(), hashlib.sha256).digest()
    return f"{header}.{payload_b64}.{_b64url(sig)}"

def jwt_decode(token: str, secret: str) -> dict:
    """Verify and decode a JWT. Raises ValueError on bad/expired token."""
    try:
        header_b64, payload_b64, sig_b64 = token.split('.')
    except ValueError:
        raise ValueError("Malformed token")
    expected = hmac.new(secret.encode(), f"{header_b64}.{payload_b64}".encode(), hashlib.sha256).digest()
    if not hmac.compare_digest(expected, _b64url_decode(sig_b64)):
        raise ValueError("Invalid signature")
    payload = json.loads(_b64url_decode(payload_b64))
    if payload.get('exp', 0) < int(time.time()):
        raise ValueError("Token expired")
    return payload

# ─────────────────────────────────────────────
# APP SETUP
# ─────────────────────────────────────────────
app = Flask(__name__)
socketio = SocketIO(app, cors_allowed_origins="*")

app.secret_key = 'SOES_ultra_secure_2026_key_#Xm9!'
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(minutes=15)
JWT_SECRET = 'SOES_jwt_secret_key_2026_#Zm7!'

# ─────────────────────────────────────────────
# STORAGE HELPERS — swap cloud backend here only
# ─────────────────────────────────────────────
RECORDINGS_DIR  = os.path.join(os.path.dirname(os.path.abspath(__file__)), "recordings")
EXAM_PAPERS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "exam_papers")
os.makedirs(RECORDINGS_DIR,  exist_ok=True)
os.makedirs(EXAM_PAPERS_DIR, exist_ok=True)

def save_recording_file(token: str, file_storage) -> str:
    """Save a recording and return its filename. Swap body for S3/GCS to move cloud."""
    d = os.path.join(RECORDINGS_DIR, token); os.makedirs(d, exist_ok=True)
    name = secure_filename(file_storage.filename) or f"meet_{token}_{int(time.time())}.webm"
    file_storage.save(os.path.join(d, name))
    return name

def save_exam_pdf_file(token: str, exam_id: str, file_storage) -> str:
    """Save an exam PDF and return its filename. Swap body for S3/GCS to move cloud."""
    d = os.path.join(EXAM_PAPERS_DIR, token); os.makedirs(d, exist_ok=True)
    name = secure_filename(f"{exam_id}_{file_storage.filename}")
    file_storage.save(os.path.join(d, name))
    return name

# ─────────────────────────────────────────────
# PERSISTENT DATA
# ─────────────────────────────────────────────
DATA_FILE = "classroom_storage.json"

# In-memory stores (overwritten by load_system_data below)
USER_DATABASE        = {}   # email -> {full_name, password_hash, role, otp}
CLASSROOM_DATABASE   = []   # [{name, token, owner_email}]
STUDENT_ENROLLMENTS  = {}   # email -> [token, ...]
EXAM_DATABASE        = {}   # token -> {upcoming, assigned_exams, assigned_assignments, deadlines}
RESULT_DATABASE      = {}
ACTIVITY_LOGS        = {}   # token -> [{timestamp, log}]
RECORDINGS_DATABASE  = {}   # token -> [{filename, timestamp, shared}]
EXAM_PAPERS_DATABASE = {}   # token -> [exam_entry]
EXAM_SUBMISSIONS_DB  = {}   # student_email -> {exam_id -> submission}
LOGIN_HISTORY        = {}   # email -> [{event, timestamp, ip, ua}]
BROWSER_EVENTS       = {}   # student_email -> [{type, detail, timestamp, exam_id}]

def load_system_data():
    global USER_DATABASE, CLASSROOM_DATABASE, STUDENT_ENROLLMENTS, EXAM_DATABASE
    global RESULT_DATABASE, ACTIVITY_LOGS, RECORDINGS_DATABASE, EXAM_PAPERS_DATABASE
    global EXAM_SUBMISSIONS_DB, LOGIN_HISTORY, BROWSER_EVENTS
    if os.path.exists(DATA_FILE):
        try:
            with open(DATA_FILE) as f:
                d = json.load(f)
            USER_DATABASE        = d.get("users", {})
            CLASSROOM_DATABASE   = d.get("classrooms", [])
            STUDENT_ENROLLMENTS  = d.get("enrollments", {})
            EXAM_DATABASE        = d.get("exams", {})
            RESULT_DATABASE      = d.get("results", {})
            ACTIVITY_LOGS        = d.get("activity_logs", {})
            RECORDINGS_DATABASE  = d.get("recordings", {})
            EXAM_PAPERS_DATABASE = d.get("exam_papers", {})
            EXAM_SUBMISSIONS_DB  = d.get("exam_submissions", {})
            LOGIN_HISTORY        = d.get("login_history", {})
            BROWSER_EVENTS       = d.get("browser_events", {})
        except Exception as e:
            print(f"[WARN] Could not load data: {e}")

def save_system_data():
    with open(DATA_FILE, "w") as f:
        json.dump({
            "users":            USER_DATABASE,
            "classrooms":       CLASSROOM_DATABASE,
            "enrollments":      STUDENT_ENROLLMENTS,
            "exams":            EXAM_DATABASE,
            "results":          RESULT_DATABASE,
            "activity_logs":    ACTIVITY_LOGS,
            "recordings":       RECORDINGS_DATABASE,
            "exam_papers":      EXAM_PAPERS_DATABASE,
            "exam_submissions": EXAM_SUBMISSIONS_DB,
            "login_history":    LOGIN_HISTORY,
            "browser_events":   BROWSER_EVENTS,
        }, f, indent=2)

load_system_data()

# ─────────────────────────────────────────────
# FORENSIC HELPERS
# ─────────────────────────────────────────────
def _now() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

def log_login_event(email: str, event: str):
    """Record login / logout events with IP and user-agent."""
    LOGIN_HISTORY.setdefault(email, []).append({
        "event":     event,
        "timestamp": _now(),
        "ip":        request.remote_addr or "unknown",
        "ua":        request.user_agent.string[:120]
    })
    save_system_data()

def log_activity(token: str, message: str):
    ACTIVITY_LOGS.setdefault(token.upper(), []).append({
        "timestamp": _now(), "log": message
    })
    save_system_data()

def log_browser_event(email: str, event_type: str, detail: str, exam_id: str = ""):
    """Store browser-monitoring records (tab switch, fullscreen exit, etc.)."""
    BROWSER_EVENTS.setdefault(email, []).append({
        "type":      event_type,
        "detail":    detail,
        "exam_id":   exam_id,
        "timestamp": _now()
    })
    save_system_data()

# ==========================================================
# ANALYTICS & REPORT HELPERS
# ==========================================================

from collections import Counter
from datetime import datetime


def calculate_student_risk(student_email, events=None):
    """
    Calculates overall cheating risk score.
    Pass `events` explicitly to score against a filtered subset
    (e.g. only violations tied to one classroom's exams).
    """

    events = BROWSER_EVENTS.get(student_email, []) if events is None else events

    score = 0

    weights = {
        "tab_switch":10,
        "fullscreen_exit":15,
        "multiple_faces":30,
        "face_missing":20,
        "head_turn_left":8,
        "head_turn_right":8,
        "looking_down":6,
        "window_blur":10
    }

    for e in events:
        score += weights.get(e["type"],2)

    if score > 100:
        score = 100

    if score < 25:
        level="LOW"

    elif score<60:
        level="MEDIUM"

    else:
        level="HIGH"

    return {
        "score":score,
        "level":level
    }


def get_login_statistics():

    today=datetime.now().date()

    daily=0
    weekly=0
    monthly=0

    for logs in LOGIN_HISTORY.values():

        for log in logs:

            try:
                d=datetime.strptime(
                    log["timestamp"],
                    "%Y-%m-%d %H:%M:%S"
                ).date()

                if d==today:
                    daily+=1

                if (today-d).days<=7:
                    weekly+=1

                if d.month==today.month and d.year==today.year:
                    monthly+=1

            except:
                pass

    return {

        "total_users":len(LOGIN_HISTORY),

        "daily":daily,

        "weekly":weekly,

        "monthly":monthly

    }


def get_submission_statistics():

    submitted=0
    late=0
    pending=0

    for student in EXAM_SUBMISSIONS_DB.values():

        for sub in student.values():

            submitted+=1

            if sub.get("late"):
                late+=1

    total_students=len(USER_DATABASE)

    pending=max(0,total_students-submitted)

    return{

        "submitted":submitted,

        "late":late,

        "pending":pending

    }


def get_defaulters():

    data=[]

    for email,user in USER_DATABASE.items():

        if user.get("role")!="student":
            continue

        risk=calculate_student_risk(email)

        violations=len(BROWSER_EVENTS.get(email,[]))

        data.append({

            "student":user["full_name"],

            "email":email,

            "violations":violations,

            "risk":risk["level"],

            "score":risk["score"]

        })

    data.sort(
        key=lambda x:x["score"],
        reverse=True
    )

    return data


def cheating_statistics():

    counter=Counter()

    for logs in BROWSER_EVENTS.values():

        for item in logs:

            counter[item["type"]]+=1

    return dict(counter)


def classroom_statistics():

    report=[]

    for room in CLASSROOM_DATABASE:

        token=room["token"]

        students=sum(
            token in cls
            for cls in STUDENT_ENROLLMENTS.values()
        )

        exams=len(
            EXAM_PAPERS_DATABASE.get(token,[])
        )

        report.append({

            "classroom":room["name"],

            "token":token,

            "students":students,

            "papers":exams,

            "activities":len(
                ACTIVITY_LOGS.get(token,[])
            )

        })

    return report

# ==========================================================
# CLASSROOM-SCOPED REPORT BUILDERS (used by the Reports tab)
# ==========================================================

def _classroom_enrolled_students(token):
    tok = token.upper()
    return [e for e, toks in STUDENT_ENROLLMENTS.items() if tok in toks]

def _classroom_exam_ids(token):
    tok = token.upper()
    return {e['id'] for e in EXAM_PAPERS_DATABASE.get(tok, [])}

def _classroom_events_by_student(token):
    """Browser/proctoring events for this classroom's students, restricted
    to violations that occurred during one of this classroom's exams."""
    exam_ids = _classroom_exam_ids(token)
    out = {}
    for email in _classroom_enrolled_students(token):
        out[email] = [e for e in BROWSER_EVENTS.get(email, []) if e.get('exam_id') in exam_ids]
    return out

def get_classroom_overview(token):
    tok = token.upper()
    ed  = EXAM_DATABASE.get(tok, {})
    return {
        "students":      len(_classroom_enrolled_students(tok)),
        "exams":         len(EXAM_PAPERS_DATABASE.get(tok, [])),
        "activities":    len(ACTIVITY_LOGS.get(tok, [])),
        "assignments":   len(ed.get('assigned_assignments', [])),
        "deadlines":     len(ed.get('deadlines', [])),
        "announcements": len(ed.get('announcements', [])),
        "recordings":    len(RECORDINGS_DATABASE.get(tok, [])),
    }

def get_classroom_submission_stats(token):
    tok      = token.upper()
    students = _classroom_enrolled_students(tok)
    exam_ids = _classroom_exam_ids(tok)
    submitted = late = 0
    for email in students:
        for eid, sub in EXAM_SUBMISSIONS_DB.get(email, {}).items():
            if eid in exam_ids:
                submitted += 1
                if sub.get('late'): late += 1
    total_possible = len(students) * len(exam_ids)
    pending = max(0, total_possible - submitted)
    return {"submitted": submitted, "late": late, "pending": pending}

def get_classroom_login_stats(token):
    today = datetime.now().date()
    daily = weekly = monthly = 0
    for email in _classroom_enrolled_students(token):
        for log in LOGIN_HISTORY.get(email, []):
            try:
                d = datetime.strptime(log["timestamp"], "%Y-%m-%d %H:%M:%S").date()
            except Exception:
                continue
            if d == today: daily += 1
            if (today - d).days <= 7: weekly += 1
            if d.month == today.month and d.year == today.year: monthly += 1
    return {"daily": daily, "weekly": weekly, "monthly": monthly}

def get_classroom_cheating_breakdown(token):
    counter = Counter()
    for events in _classroom_events_by_student(token).values():
        for item in events:
            counter[item["type"]] += 1
    return dict(counter)

def get_classroom_defaulters(token):
    data = []
    events_by_student = _classroom_events_by_student(token)
    for email in _classroom_enrolled_students(token):
        user   = USER_DATABASE.get(email, {})
        events = events_by_student.get(email, [])
        risk   = calculate_student_risk(email, events=events)
        data.append({
            "student":    user.get("full_name", email.split('@')[0]),
            "email":      email,
            "violations": len(events),
            "risk":       risk["level"],
            "score":      risk["score"],
        })
    data.sort(key=lambda x: x["score"], reverse=True)
    return data

def build_classroom_report(token):
    tok = token.upper()
    return {
        "overview":    get_classroom_overview(tok),
        "submissions": get_classroom_submission_stats(tok),
        "logins":      get_classroom_login_stats(tok),
        "cheating":    get_classroom_cheating_breakdown(tok),
        "defaulters":  get_classroom_defaulters(tok),
        "generated_at": _now(),
    }

# ─────────────────────────────────────────────
# VALIDATION HELPERS
# ─────────────────────────────────────────────
def validate_password(pw: str) -> bool:
    if not pw or len(pw) > 20: return False
    if not re.search(r"[A-Z]", pw): return False
    if not re.search(r"[0-9]", pw): return False
    if not re.search(r'[!@#$%^&*(),.?":{}|<>]', pw): return False
    return True

def owned_classroom(token: str, admin_email: str):
    c = next((x for x in CLASSROOM_DATABASE if x['token'] == token.upper()), None)
    return c if (c and c.get('owner_email') == admin_email) else None

def _new_exam_id() -> str:
    import uuid; return uuid.uuid4().hex[:10]

def _unique_token() -> str:
    existing = {c['token'] for c in CLASSROOM_DATABASE}
    chars = "ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789"
    while True:
        t = ''.join(random.choice(chars) for _ in range(6))
        if t not in existing: return t

# ─────────────────────────────────────────────
# EXAM PANEL SYNC
# Keeps upcoming / live / deadlines in sync with EXAM_PAPERS_DATABASE
# ─────────────────────────────────────────────
def sync_exam_panels(token: str):
    papers = EXAM_PAPERS_DATABASE.get(token.upper(), [])
    now    = datetime.now()
    upcoming, live, closed = [], [], []
    for p in papers:
        try:
            start = datetime.fromisoformat(p['start_time'])
        except Exception:
            continue
        end   = start + timedelta(minutes=int(p.get('duration_minutes', 60)))
        label = f"{p['title']}  ·  {p['start_time'].replace('T',' ')}  ·  {p['duration_minutes']} min  ·  {p['type'].upper()}"
        if now < start:
            upcoming.append(label)
        elif now <= end:
            live.append(label + "  🟢 LIVE NOW")
        else:
            closed.append(label + "  ✓ Closed")
    existing = EXAM_DATABASE.get(token.upper(), {})
    EXAM_DATABASE[token.upper()] = {
        "upcoming":              upcoming,
        "assigned_exams":        live,
        "assigned_assignments":  existing.get("assigned_assignments", []),
        "deadlines":             closed,
        "announcements":         existing.get("announcements", [])
    }

# ─────────────────────────────────────────────
# SESSION & JWT DECORATORS
# ─────────────────────────────────────────────
@app.before_request
def _session_permanent():
    session.permanent = True

def require_role(role: str):
    def decorator(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            if session.get('user_type') != role:
                return redirect(url_for('admin_login' if role == 'admin' else 'student_login'))
            # Session monitoring: check JWT token validity too
            token = session.get('jwt_token', '')
            try:
                jwt_decode(token, JWT_SECRET)
            except ValueError:
                session.clear()
                return redirect(url_for('admin_login' if role == 'admin' else 'student_login'))
            return f(*args, **kwargs)
        return wrapper
    return decorator

# ─────────────────────────────────────────────
# DEVICE WARNING (server-side mobile detection)
# ─────────────────────────────────────────────
MOBILE_UA = re.compile(r"(Mobile|Android|iPhone|iPad|iPod|BlackBerry|Windows Phone)", re.I)

@app.before_request
def _device_check():
    # Pages that require desktop
    protected = ['/admin-login', '/student-login', '/admin-dashboard',
                 '/student-dashboard', '/classroom/', '/student/classroom/',
                 '/exam/', '/meet/']
    path = request.path
    if any(path.startswith(p) for p in protected):
        ua = request.user_agent.string
        if MOBILE_UA.search(ua) and not session.get('device_warned'):
            session['device_warning_pending'] = True

# ─────────────────────────────────────────────
# HOME
# ─────────────────────────────────────────────
@app.route('/')
def home():
    return render_template('home.html')

# ─────────────────────────────────────────────
# SIGNUP
# ─────────────────────────────────────────────
def _create_account(email, full_name, password, role):
    if not email or not email.endswith('@gmail.com'):
        return False, "Please use a valid @gmail.com address."
    if email in USER_DATABASE:
        return False, "An account with this email already exists. Please log in instead."
    if not full_name or not full_name.strip():
        return False, "Full name is required."
    if not validate_password(password):
        return False, ("Password must be ≤ 20 characters and include an uppercase letter, "
                       "a number, and a special character.")
    USER_DATABASE[email] = {
        "full_name":  full_name.strip(),
        "password":   generate_password_hash(password),
        "role":       role,
        "otp":        None
    }
    save_system_data()
    return True, None

@app.route('/admin-signup', methods=['GET', 'POST'])
def admin_signup():
    if request.method == 'POST':
        ok, err = _create_account(request.form.get('identity'),
                                   request.form.get('full_name'),
                                   request.form.get('password'), 'admin')
        if not ok:
            return render_template('admin_signup.html', error=err,
                                   full_name=request.form.get('full_name'),
                                   identity=request.form.get('identity'))
        email = request.form.get('identity')
        token = jwt_encode({'email': email, 'role': 'admin'}, JWT_SECRET)
        session.update({'user_type': 'admin', 'user_email': email, 'jwt_token': token})
        log_login_event(email, 'signup+login')
        return redirect(url_for('admin_dashboard'))
    return render_template('admin_signup.html')

@app.route('/student-signup', methods=['GET', 'POST'])
def student_signup():
    if request.method == 'POST':
        ok, err = _create_account(request.form.get('identity'),
                                   request.form.get('full_name'),
                                   request.form.get('password'), 'student')
        if not ok:
            return render_template('student_signup.html', error=err,
                                   full_name=request.form.get('full_name'),
                                   identity=request.form.get('identity'))
        email = request.form.get('identity')
        token = jwt_encode({'email': email, 'role': 'student'}, JWT_SECRET)
        session.update({'user_type': 'student', 'user_email': email, 'jwt_token': token})
        log_login_event(email, 'signup+login')
        return redirect(url_for('student_dashboard'))
    return render_template('student_signup.html')

# ─────────────────────────────────────────────
# LOGIN / LOGOUT
# ─────────────────────────────────────────────
@app.route('/send-otp', methods=['POST'])
def send_otp():
    email = request.form.get('email')
    if email not in USER_DATABASE:
        return jsonify({"success": False, "message": "Email not found."}), 404
    otp_code = str(random.randint(100000, 999999))
    USER_DATABASE[email]['otp'] = otp_code
    save_system_data()
    print(f"\n{'='*40}\n OTP for {email}: {otp_code}\n{'='*40}\n")
    return jsonify({"success": True, "message": "OTP sent to console."})

def _do_login(email, password, otp_attempt, login_mode, role, login_tpl, dashboard_fn):
    user = USER_DATABASE.get(email)
    if not user or not email.endswith('@gmail.com'):
        return render_template(login_tpl, error="No account found. Please sign up first."), 401
    if user.get('role') != role:
        return render_template(login_tpl, error=f"This account is not registered as a {role}."), 401

    ok = ((login_mode == 'otp'      and otp_attempt and otp_attempt == user['otp']) or
          (login_mode == 'password' and password    and check_password_hash(user['password'], password)))
    if not ok:
        log_login_event(email, 'failed_login')
        return render_template(login_tpl, error="Incorrect credentials."), 401

    if login_mode == 'otp': user['otp'] = None
    save_system_data()
    jwt_tok = jwt_encode({'email': email, 'role': role}, JWT_SECRET)
    session.update({'user_type': role, 'user_email': email, 'jwt_token': jwt_tok})
    log_login_event(email, 'login')
    return redirect(url_for(dashboard_fn))

@app.route('/admin-login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        return _do_login(request.form.get('identity'), request.form.get('password'),
                         request.form.get('otp'), request.form.get('login_mode'),
                         'admin', 'admin.html', 'admin_dashboard')
    return render_template('admin.html')

@app.route('/student-login', methods=['GET', 'POST'])
def student_login():
    if request.method == 'POST':
        return _do_login(request.form.get('identity'), request.form.get('password'),
                         request.form.get('otp'), request.form.get('login_mode'),
                         'student', 'student.html', 'student_dashboard')
    return render_template('student.html')

@app.route('/logout')
def logout():
    email = session.get('user_email', '')
    if email: log_login_event(email, 'logout')
    session.clear()
    return redirect(url_for('home'))

# ─────────────────────────────────────────────
# DEVICE WARNING ACKNOWLEDGE
# ─────────────────────────────────────────────
@app.route('/acknowledge-device-warning', methods=['POST'])
def ack_device_warning():
    session['device_warned'] = True
    session.pop('device_warning_pending', None)
    return jsonify({"ok": True})

# ─────────────────────────────────────────────
# ADMIN DASHBOARD
# ─────────────────────────────────────────────
@app.route('/admin-dashboard')
@require_role('admin')
def admin_dashboard():
    email  = session['user_email']
    name   = USER_DATABASE.get(email, {}).get('full_name', 'Administrator')
    rooms  = [c for c in CLASSROOM_DATABASE if c.get('owner_email') == email]
    warn   = session.pop('device_warning_pending', False)
    return render_template('admin_dashboard.html', admin_name=name,
                           classrooms=rooms, device_warning=warn)

@app.route('/create-classroom', methods=['POST'])
@require_role('admin')
def create_classroom():
    name  = request.form.get('class_name', '').strip()
    email = session['user_email']
    if name:
        tok = _unique_token()
        CLASSROOM_DATABASE.append({"name": name, "token": tok, "owner_email": email})
        save_system_data()
        log_activity(tok, f"Classroom created by {email}")
    return redirect(url_for('admin_dashboard'))

@app.route('/delete-classroom/<token>', methods=['POST'])
@require_role('admin')
def delete_classroom(token):
    global CLASSROOM_DATABASE
    email = session['user_email']
    t     = token.upper()
    c     = next((x for x in CLASSROOM_DATABASE if x['token'] == t), None)
    if not c or c.get('owner_email') != email:
        return "Access Denied", 403
    CLASSROOM_DATABASE = [x for x in CLASSROOM_DATABASE if x['token'] != t]
    for s in STUDENT_ENROLLMENTS:
        if t in STUDENT_ENROLLMENTS[s]: STUDENT_ENROLLMENTS[s].remove(t)
    save_system_data()
    return redirect(url_for('admin_dashboard'))

# ─────────────────────────────────────────────
# ADMIN CLASSROOM
# ─────────────────────────────────────────────
@app.route('/classroom/<token>')
@require_role('admin')
def classroom_home(token):
    email     = session['user_email']
    classroom = owned_classroom(token, email)
    if not classroom: return "Access Denied or Not Found", 403
    tok = token.upper()
    sync_exam_panels(tok)
    exams_data  = EXAM_DATABASE.get(tok, {"upcoming":[],"assigned_exams":[],"assigned_assignments":[],"deadlines":[],"announcements":[]})
    pdf_bank    = EXAM_DATABASE.get(f"{tok}_PDFS", [])
    activities  = ACTIVITY_LOGS.get(tok, [])
    recordings  = RECORDINGS_DATABASE.get(tok, [])
    exam_papers = EXAM_PAPERS_DATABASE.get(tok, [])
    # Build per-student submission counts for admin view
    enrolled    = [e for e, toks in STUDENT_ENROLLMENTS.items() if tok in toks]
    return render_template('classroom.html', classroom=classroom,
                           exams=exams_data, pdf_bank=pdf_bank,
                           activities=activities, recordings=recordings,
                           exam_papers=exam_papers, enrolled=enrolled)

@app.route('/classroom/<token>/post-assignment', methods=['POST'])
@require_role('admin')
def post_assignment(token):
    tok  = token.upper()
    text = request.form.get('assignment_text', '').strip()
    dl   = request.form.get('deadline', '').strip()
    kind = request.form.get('kind', 'assignment')
    if not text: return redirect(url_for('classroom_home', token=tok))
    EXAM_DATABASE.setdefault(tok, {"upcoming":[],"assigned_exams":[],"assigned_assignments":[],"deadlines":[],"announcements":[]})
    if kind == 'announcement':
        # Announcements are stored as structured entries so the board can be
        # rendered richly (author, posted time, optional target date) and
        # kept in sync with the student-facing classroom.
        EXAM_DATABASE[tok].setdefault('announcements', []).append({
            "text":      text,
            "deadline":  dl,
            "posted_at": _now(),
            "author":    USER_DATABASE.get(session['user_email'], {}).get('full_name', 'Admin')
        })
        log_activity(tok, f"Admin posted a public announcement: \"{text[:80]}\"")
    else:
        label = text + (f"  |  Due: {dl}" if dl else "")
        if kind == 'deadline':
            EXAM_DATABASE[tok].setdefault('deadlines', []).append(label)
            log_activity(tok, f"Admin added a deadline alert: \"{text[:80]}\"")
        else:
            EXAM_DATABASE[tok].setdefault('assigned_assignments', []).append(label)
            log_activity(tok, f"Admin posted a new assignment task: \"{text[:80]}\"")
    save_system_data()
    return redirect(url_for('classroom_home', token=tok))

@app.route('/classroom/<token>/delete-assignment', methods=['POST'])
@require_role('admin')
def delete_assignment(token):
    tok   = token.upper()
    kind  = request.form.get('kind', 'assignment')
    index = int(request.form.get('index', -1))
    if tok in EXAM_DATABASE:
        key   = {'deadline': 'deadlines', 'announcement': 'announcements'}.get(kind, 'assigned_assignments')
        items = EXAM_DATABASE[tok].get(key, [])
        if 0 <= index < len(items): items.pop(index)
    save_system_data()
    return redirect(url_for('classroom_home', token=tok))

@app.route('/classroom/<token>/activity-logs.json')
@require_role('admin')
def activity_logs_json(token):
    """Polled by the Activity Logs panel to stream in new entries live,
    and to pick up anything logged since midnight on a fresh day."""
    if not owned_classroom(token, session['user_email']): return jsonify({"success": False}), 403
    tok = token.upper()
    since = request.args.get('since', '')  # ISO-ish "YYYY-MM-DD HH:MM:SS" of the last entry the client has
    activities = ACTIVITY_LOGS.get(tok, [])
    if since:
        activities = [a for a in activities if a["timestamp"] > since]
    return jsonify({"success": True, "activities": activities, "total": len(ACTIVITY_LOGS.get(tok, []))})

# ─────────────────────────────────────────────
# FORENSIC: ADMIN LOGS VIEWS
# ─────────────────────────────────────────────
@app.route('/admin/forensics')
@require_role('admin')
def forensics_dashboard():
    email = session['user_email']
    # Show login history for students in admin's classrooms
    my_tokens    = {c['token'] for c in CLASSROOM_DATABASE if c.get('owner_email') == email}
    enrolled_students = [e for e, toks in STUDENT_ENROLLMENTS.items()
                         if any(t in my_tokens for t in toks)]
    student_logins = {s: LOGIN_HISTORY.get(s, []) for s in enrolled_students}
    student_events = {s: BROWSER_EVENTS.get(s, [])   for s in enrolled_students}
    admin_logins   = LOGIN_HISTORY.get(email, [])
    activity_logs  = {t: ACTIVITY_LOGS.get(t, []) for t in my_tokens}
    return render_template('forensics.html',
                           admin_logins=admin_logins,
                           student_logins=student_logins,
                           student_events=student_events,
                           activity_logs=activity_logs,
                           my_tokens=my_tokens)

# ─────────────────────────────────────────────
# EXAM PAPER MANAGEMENT
# ─────────────────────────────────────────────
@app.route('/classroom/<token>/create-exam', methods=['POST'])
@require_role('admin')
def create_exam(token):
    if not owned_classroom(token, session['user_email']):
        return "Access Denied", 403
    tok       = token.upper()
    exam_type = request.form.get('exam_type')
    title     = request.form.get('title', '').strip()
    duration  = int(request.form.get('duration_minutes', 60))
    start_time = request.form.get('start_time')
    if not title or not start_time:
        return "Title and start time are required.", 400
    entry = {"id": _new_exam_id(), "title": title, "type": exam_type,
             "duration_minutes": duration, "start_time": start_time,
             "created_at": _now()}
    if exam_type == 'pdf':
        f = request.files.get('exam_pdf')
        if not f or f.filename == '': return "PDF required.", 400
        entry["pdf_filename"] = save_exam_pdf_file(tok, entry['id'], f)
    elif exam_type == 'mcq':
        qt   = request.form.getlist('question_text[]')
        opts = [request.form.getlist(f'option_{i}[]') for i in range(4)]
        corr = request.form.getlist('correct_index[]')
        qs   = [{"question": q.strip(), "options": [opts[o][i] for o in range(4)],
                 "correct_index": int(corr[i])}
                for i, q in enumerate(qt) if q.strip()]
        if not qs: return "At least one question required.", 400
        entry["questions"] = qs
    else:
        return "Invalid exam type.", 400
    EXAM_PAPERS_DATABASE.setdefault(tok, []).append(entry)
    sync_exam_panels(tok)
    save_system_data()
    log_activity(tok, f"Exam '{title}' created by {session['user_email']}")
    return redirect(url_for('classroom_home', token=tok))

@app.route('/classroom/<token>/delete-exam/<exam_id>', methods=['POST'])
@require_role('admin')
def delete_exam(token, exam_id):
    if not owned_classroom(token, session['user_email']): return "Access Denied", 403
    tok = token.upper()
    EXAM_PAPERS_DATABASE[tok] = [e for e in EXAM_PAPERS_DATABASE.get(tok, []) if e['id'] != exam_id]
    sync_exam_panels(tok)
    save_system_data()
    return redirect(url_for('classroom_home', token=tok))

# ─────────────────────────────────────────────
# REPORTS TAB
# ─────────────────────────────────────────────
@app.route('/classroom/<token>/reports')
@require_role('admin')
def classroom_reports(token):
    classroom = owned_classroom(token, session['user_email'])
    if not classroom: return "Access Denied", 403
    tok    = token.upper()
    report = build_classroom_report(tok)
    return render_template('reports.html', classroom=classroom, report=report,
                           token=tok, openpyxl_available=OPENPYXL_AVAILABLE)

def _rows_for_report(kind, report):
    if kind == 'overview':
        headers = ['Metric', 'Value']
        rows = [[k.replace('_', ' ').title(), v] for k, v in report['overview'].items()]
    elif kind == 'submissions':
        headers = ['Status', 'Count']
        s = report['submissions']
        rows = [['Submitted', s['submitted']], ['Late', s['late']], ['Pending', s['pending']]]
    elif kind == 'logins':
        headers = ['Period', 'Logins']
        l = report['logins']
        rows = [['Today', l['daily']], ['Last 7 Days', l['weekly']], ['This Month', l['monthly']]]
    elif kind == 'cheating':
        headers = ['Violation Type', 'Count']
        rows = [[k.replace('_', ' ').title(), v] for k, v in report['cheating'].items()]
    elif kind == 'defaulters':
        headers = ['Student', 'Email', 'Violations', 'Risk Level', 'Risk Score']
        rows = [[d['student'], d['email'], d['violations'], d['risk'], d['score']] for d in report['defaulters']]
    else:
        headers, rows = ['Error'], [['Unknown report type']]
    return headers, rows

@app.route('/classroom/<token>/reports/export/<kind>.<fmt>')
@require_role('admin')
def export_classroom_report(token, kind, fmt):
    classroom = owned_classroom(token, session['user_email'])
    if not classroom: return "Access Denied", 403
    tok            = token.upper()
    report         = build_classroom_report(tok)
    headers, rows  = _rows_for_report(kind, report)
    safe_name      = re.sub(r'[^A-Za-z0-9_-]+', '_', classroom['name'])
    fname_base     = f"{safe_name}_{kind}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

    if fmt == 'csv':
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(headers)
        writer.writerows(rows)
        return Response(buf.getvalue(), mimetype='text/csv',
                         headers={'Content-Disposition': f'attachment; filename="{fname_base}.csv"'})
    elif fmt == 'xlsx':
        if not OPENPYXL_AVAILABLE:
            return ("Excel export requires the 'openpyxl' package on the server. "
                    "Install it with: pip install openpyxl — or use the CSV export instead.", 501)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = kind.title()[:31]
        ws.append(headers)
        for r in rows: ws.append(r)
        for i, h in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(i)].width = max(14, len(str(h)) + 4)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, as_attachment=True, download_name=f"{fname_base}.xlsx",
                          mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    else:
        return "Unsupported export format", 400

# ─────────────────────────────────────────────
# SECURE QUESTION DELIVERY  (JWT-gated PDF serve)
# ─────────────────────────────────────────────
@app.route('/exam-token/<token>/<exam_id>')
def get_exam_token(token, exam_id):
    """Issue a short-lived JWT that unlocks the exam PDF for one student."""
    if session.get('user_type') != 'student': return "Forbidden", 403
    email = session.get('user_email', '')
    tok   = token.upper()
    if tok not in STUDENT_ENROLLMENTS.get(email, []):
        return "Not enrolled", 403
    paper = next((e for e in EXAM_PAPERS_DATABASE.get(tok, []) if e['id'] == exam_id), None)
    if not paper: return "Exam not found", 404
    start = datetime.fromisoformat(paper['start_time'])
    if datetime.now() < start:
        return "Exam has not started yet", 403
    access_token = jwt_encode({'email': email, 'token': tok, 'exam_id': exam_id}, JWT_SECRET, expires_in=7200)
    return jsonify({"access_token": access_token})

@app.route('/exam-paper-file/<token>/<exam_id>')
def serve_exam_pdf(token, exam_id):
    """Serve the exam PDF only to JWT-verified students or the owning admin."""
    tok   = token.upper()
    paper = next((e for e in EXAM_PAPERS_DATABASE.get(tok, []) if e['id'] == exam_id), None)
    if not paper or paper.get('type') != 'pdf': return "Not found", 404

    if session.get('user_type') == 'admin':
        if not owned_classroom(tok, session.get('user_email', '')): return "Forbidden", 403
    elif session.get('user_type') == 'student':
        # Verify the JWT access token from query param
        access_token = request.args.get('t', '')
        try:
            payload = jwt_decode(access_token, JWT_SECRET)
            assert payload['exam_id'] == exam_id and payload['token'] == tok
        except Exception:
            return "Access denied — obtain an exam token first via /exam-token/.", 403
    else:
        return "Forbidden", 403
    return send_from_directory(os.path.join(EXAM_PAPERS_DIR, tok), paper['pdf_filename'])

# ─────────────────────────────────────────────
# EXAM ROOM (student)
# ─────────────────────────────────────────────
@app.route('/exam/<token>/<exam_id>')
@require_role('student')
def take_exam(token, exam_id):
    email = session['user_email']
    tok   = token.upper()
    if tok not in STUDENT_ENROLLMENTS.get(email, []):
        return "Not enrolled", 403
    paper = next((e for e in EXAM_PAPERS_DATABASE.get(tok, []) if e['id'] == exam_id), None)
    if not paper: return "Exam not found", 404
    if exam_id in EXAM_SUBMISSIONS_DB.get(email, {}):
        return "Already submitted", 403
    classroom = next((c for c in CLASSROOM_DATABASE if c['token'] == tok), None)
    # Issue an exam JWT so the PDF can be securely fetched
    access_token = jwt_encode({'email': email, 'token': tok, 'exam_id': exam_id}, JWT_SECRET, expires_in=7200)
    log_activity(tok, f"Student {email} entered exam '{paper['title']}'")
    return render_template('exam_room.html', paper=paper, token=tok,
                           classroom=classroom, student_email=email,
                           access_token=access_token)

@app.route('/exam/<token>/<exam_id>/submit', methods=['POST'])
@require_role('student')
def submit_exam(token, exam_id):
    email = session['user_email']
    tok   = token.upper()
    paper = next((e for e in EXAM_PAPERS_DATABASE.get(tok, []) if e['id'] == exam_id), None)
    if not paper: return jsonify({"success": False}), 404
    payload    = request.get_json(force=True) or {}
    answers    = payload.get('answers', {})
    violations = payload.get('violations', [])
    score = total = None
    if paper.get('type') == 'mcq':
        qs    = paper.get('questions', [])
        total = len(qs)
        score = sum(1 for i, q in enumerate(qs)
                    if answers.get(str(i)) is not None
                    and int(answers[str(i)]) == q['correct_index'])
    EXAM_SUBMISSIONS_DB.setdefault(email, {})[exam_id] = {
        "exam_title": paper['title'], "token": tok,
        "answers": answers, "score": score, "total": total,
        "violations": violations, "submitted_at": _now()
    }
    # Log every violation as a browser event
    for v in violations:
        log_browser_event(email, v.get('type', 'violation'), v.get('reason', ''), exam_id)
    save_system_data()
    log_activity(tok, f"Student {email} submitted exam '{paper['title']}' — score {score}/{total}")
    return jsonify({"success": True, "score": score, "total": total})

# ─────────────────────────────────────────────
# LIVE PROCTORING (browser events → server)
# ─────────────────────────────────────────────
@app.route('/exam/<token>/<exam_id>/flag', methods=['POST'])
@require_role('student')
def flag_violation(token, exam_id):
    email  = session['user_email']
    tok    = token.upper()
    data   = request.get_json(force=True) or {}
    reason = data.get('reason', 'Unspecified')
    # Store in forensic browser events log
    log_browser_event(email, 'proctoring_flag', reason, exam_id)
    # Push live to admin monitor
    socketio.emit('proctor-alert',
                  {"student": email, "reason": reason, "timestamp": _now()},
                  room=f"proctor-{tok}-{exam_id}")
    return jsonify({"success": True})

@app.route('/browser-event', methods=['POST'])
@require_role('student')
def browser_event():
    """Endpoint for general browser-monitoring records (tab switch, visibility, etc.)"""
    email   = session['user_email']
    data    = request.get_json(force=True) or {}
    log_browser_event(email, data.get('type',''), data.get('detail',''), data.get('exam_id',''))
    return jsonify({"ok": True})

@app.route('/classroom/<token>/proctor/<exam_id>')
@require_role('admin')
def proctor_monitor(token, exam_id):
    classroom = owned_classroom(token, session['user_email'])
    if not classroom: return "Access Denied", 403
    tok   = token.upper()
    paper = next((e for e in EXAM_PAPERS_DATABASE.get(tok, []) if e['id'] == exam_id), None)
    if not paper: return "Not found", 404
    return render_template('proctor_monitor.html', paper=paper, token=tok, classroom=classroom)

# ─────────────────────────────────────────────
# STUDENT DASHBOARD & CLASSROOM
# ─────────────────────────────────────────────
@app.route('/student-dashboard')
@require_role('student')
def student_dashboard():
    email       = session['user_email']
    name        = USER_DATABASE.get(email, {}).get('full_name', email.split('@')[0])
    joined      = STUDENT_ENROLLMENTS.get(email, [])
    my_classes  = [c for c in CLASSROOM_DATABASE if c['token'] in joined]
    warn        = session.pop('device_warning_pending', False)
    return render_template('student_gateway.html', student_email=email,
                           student_name=name, my_classes=my_classes, device_warning=warn)

@app.route('/join-classroom', methods=['POST'])
@require_role('student')
def join_classroom():
    email = session['user_email']
    code  = request.form.get('class_code', '').strip().upper()
    if not next((c for c in CLASSROOM_DATABASE if c['token'] == code), None):
        return redirect(url_for('student_dashboard', error="Invalid token."))
    STUDENT_ENROLLMENTS.setdefault(email, [])
    if code not in STUDENT_ENROLLMENTS[email]:
        STUDENT_ENROLLMENTS[email].append(code)
        save_system_data()
        log_activity(code, f"Student {email} joined the classroom")
    return redirect(url_for('student_dashboard'))

@app.route('/student/classroom/<token>')
@require_role('student')
def student_classroom_home(token):
    email = session['user_email']
    tok   = token.upper()
    if tok not in STUDENT_ENROLLMENTS.get(email, []):
        return "Not enrolled", 403
    classroom = next((c for c in CLASSROOM_DATABASE if c['token'] == tok), None)
    if not classroom: return "Not found", 404
    name  = USER_DATABASE.get(email, {}).get('full_name', email.split('@')[0])
    sync_exam_panels(tok)
    exams_data  = EXAM_DATABASE.get(tok, {"upcoming":[],"assigned_exams":[],"assigned_assignments":[],"deadlines":[],"announcements":[]})
    pdf_bank    = EXAM_DATABASE.get(f"{tok}_PDFS", [])
    recordings  = [r for r in RECORDINGS_DATABASE.get(tok, []) if r.get('shared')]
    exam_papers = EXAM_PAPERS_DATABASE.get(tok, [])
    subs        = EXAM_SUBMISSIONS_DB.get(email, {})
    for p in exam_papers:
        p['_submitted'] = p['id'] in subs
    return render_template('student_classroom.html', classroom=classroom,
                           exams=exams_data, pdf_bank=pdf_bank,
                           recordings=recordings, exam_papers=exam_papers,
                           student_name=name)

# ─────────────────────────────────────────────
# RECORDINGS
# ─────────────────────────────────────────────
@app.route('/upload-recording', methods=['POST'])
@require_role('admin')
def upload_recording():
    tok  = request.form.get('token', '').strip().upper()
    file = request.files.get('recording')
    if not tok or not file: return jsonify({"success": False}), 400
    if not owned_classroom(tok, session['user_email']): return jsonify({"success": False}), 403
    name = save_recording_file(tok, file)
    RECORDINGS_DATABASE.setdefault(tok, []).append(
        {"filename": name, "timestamp": _now(), "shared": False})
    save_system_data()
    return jsonify({"success": True, "filename": name})

@app.route('/recordings/<token>/<filename>')
def serve_recording(token, filename):
    tok = token.upper()
    if session.get('user_type') == 'admin':
        if not owned_classroom(tok, session.get('user_email','')): return "Forbidden", 403
        return send_from_directory(os.path.join(RECORDINGS_DIR, tok), filename, as_attachment=True)
    if session.get('user_type') == 'student':
        email = session.get('user_email','')
        if tok not in STUDENT_ENROLLMENTS.get(email, []): return "Forbidden", 403
        rec = next((r for r in RECORDINGS_DATABASE.get(tok,[]) if r['filename']==filename), None)
        if not rec or not rec.get('shared'): return "Not shared", 403
        return send_from_directory(os.path.join(RECORDINGS_DIR, tok), filename, as_attachment=True)
    return "Forbidden", 403

@app.route('/toggle-recording-share', methods=['POST'])
@require_role('admin')
def toggle_recording_share():
    tok      = request.form.get('token','').strip().upper()
    filename = request.form.get('filename','').strip()
    if not owned_classroom(tok, session['user_email']): return jsonify({"success": False}), 403
    rec = next((r for r in RECORDINGS_DATABASE.get(tok,[]) if r['filename']==filename), None)
    if not rec: return jsonify({"success": False}), 404
    rec['shared'] = not rec.get('shared', False)
    save_system_data()
    return jsonify({"success": True, "shared": rec['shared']})

# ─────────────────────────────────────────────
# MEET ROOM
# ─────────────────────────────────────────────
@app.route('/meet/<token>')
def meet_room(token):
    if 'user_type' not in session: return redirect(url_for('home'))
    tok      = token.upper()
    is_admin = session['user_type'] == 'admin'
    if is_admin:
        classroom = owned_classroom(tok, session.get('user_email',''))
        if not classroom: return "Forbidden", 403
        self_name = USER_DATABASE.get(session['user_email'],{}).get('full_name','Admin')
        back_url  = url_for('classroom_home', token=tok)
    else:
        classroom = next((c for c in CLASSROOM_DATABASE if c['token']==tok), None)
        if not classroom: return "Not found", 404
        email = session.get('user_email','')
        if tok not in STUDENT_ENROLLMENTS.get(email,[]): return "Not enrolled", 403
        self_name = USER_DATABASE.get(email,{}).get('full_name', email.split('@')[0])
        back_url  = url_for('student_classroom_home', token=tok)
    return render_template('meet.html', classroom=classroom, is_admin=is_admin,
                           self_name=self_name, back_url=back_url)

# ─────────────────────────────────────────────
# SOCKET.IO — WebRTC signalling
# ─────────────────────────────────────────────
MEET_PARTICIPANTS = {}

@socketio.on('join-proctor-room')
def _on_join_proctor(data):
    join_room(f"proctor-{data.get('token','').upper()}-{data.get('exam_id','')}")

@socketio.on('join-room')
def _on_join_room(data):
    room = data.get('room','').upper()
    name = data.get('name','Participant')
    join_room(room)
    MEET_PARTICIPANTS[request.sid] = {"room": room, "name": name}
    emit('user-joined', {"id": request.sid, "name": name}, room=room, include_self=False)

@socketio.on('offer')
def _on_offer(d):
    if d.get('target'):
        emit('offer', {"from": request.sid, "offer": d['offer'], "name": d.get('name')}, to=d['target'])

@socketio.on('answer')
def _on_answer(d):
    if d.get('target'):
        emit('answer', {"from": request.sid, "answer": d['answer']}, to=d['target'])

@socketio.on('ice-candidate')
def _on_ice(d):
    if d.get('target'):
        emit('ice-candidate', {"from": request.sid, "candidate": d['candidate']}, to=d['target'])

@socketio.on('recording-status')
def _on_rec_status(d):
    emit('recording-status', {"recording": bool(d.get('recording'))},
         room=d.get('room','').upper(), include_self=False)

@socketio.on('leave-room')
def _on_leave(d):
    room = d.get('room','').upper()
    leave_room(room)
    MEET_PARTICIPANTS.pop(request.sid, None)
    emit('user-left', {"id": request.sid}, room=room, include_self=False)

@socketio.on('disconnect')
def _on_disconnect():
    info = MEET_PARTICIPANTS.pop(request.sid, None)
    if info:
        emit('user-left', {"id": request.sid}, room=info['room'], include_self=False)

if __name__ == '__main__':
    socketio.run(app, debug=True, port=5000)
